"""Dynamic report builder — runs user-defined queries across all entity types."""
from __future__ import annotations

import io
import uuid
from datetime import date, datetime
from typing import Any, Optional

from sqlalchemy.orm import Session, joinedload


# ---------------------------------------------------------------------------
# Entity schema definitions
# ---------------------------------------------------------------------------

def _safe(obj: Any, *attrs: str, default=None):
    """Safely navigate a chain of attributes."""
    for attr in attrs:
        if obj is None:
            return default
        obj = getattr(obj, attr, None)
    return obj if obj is not None else default


def _fmt(val: Any) -> Any:
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y %H:%M")
    if isinstance(val, date):
        return val.strftime("%d/%m/%Y")
    if isinstance(val, bool):
        return "כן" if val else "לא"
    return val


def _build_schemas():
    from app.models.elevator import Elevator
    from app.models.service_call import ServiceCall
    from app.models.customer import Customer
    from app.models.contract import Contract
    from app.models.invoice import Invoice
    from app.models.lead import Lead
    from app.models.part import Part
    from app.models.maintenance import MaintenanceSchedule
    from app.models.inspection_report import InspectionReport
    from app.models.technician import Technician

    return {
        "service_calls": {
            "label_he": "קריאות שירות",
            "model": ServiceCall,
            "eager": [joinedload(ServiceCall.elevator), joinedload(ServiceCall.assignments)],
            "default_columns": ["created_at", "address", "city", "status", "priority", "fault_type", "reported_by"],
            "columns": {
                "id":             {"label_he": "מזהה", "type": "text",     "filter": ServiceCall.id,          "get": lambda o: str(o.id)},
                "created_at":     {"label_he": "תאריך פתיחה", "type": "datetime", "filter": ServiceCall.created_at, "get": lambda o: _fmt(o.created_at)},
                "resolved_at":    {"label_he": "תאריך סגירה", "type": "datetime", "filter": ServiceCall.resolved_at, "get": lambda o: _fmt(o.resolved_at)},
                "status":         {"label_he": "סטטוס",   "type": "select", "filter": ServiceCall.status,    "get": lambda o: o.status,
                                   "options": ["OPEN","ASSIGNED","IN_PROGRESS","RESOLVED","CLOSED","MONITORING"]},
                "priority":       {"label_he": "עדיפות",  "type": "select", "filter": ServiceCall.priority,  "get": lambda o: o.priority,
                                   "options": ["CRITICAL","HIGH","MEDIUM","LOW"]},
                "fault_type":     {"label_he": "סוג תקלה","type": "select", "filter": ServiceCall.fault_type,"get": lambda o: o.fault_type,
                                   "options": ["STUCK","DOOR","ELECTRICAL","MECHANICAL","SOFTWARE","RESCUE","MAINTENANCE","OTHER"]},
                "reported_by":    {"label_he": "מדווח",   "type": "text",   "filter": ServiceCall.reported_by,"get": lambda o: o.reported_by},
                "description":    {"label_he": "תיאור",   "type": "text",   "filter": ServiceCall.description,"get": lambda o: o.description},
                "address":        {"label_he": "כתובת",   "type": "text",   "filter": Elevator.address,       "get": lambda o: _safe(o, "elevator", "address")},
                "city":           {"label_he": "עיר",     "type": "text",   "filter": Elevator.city,          "get": lambda o: _safe(o, "elevator", "city")},
                "elevator_model": {"label_he": "דגם מעלית","type": "text",  "filter": Elevator.model,         "get": lambda o: _safe(o, "elevator", "model")},
                "is_recurring":   {"label_he": "חוזרת",   "type": "bool",   "filter": ServiceCall.is_recurring,"get": lambda o: _fmt(o.is_recurring)},
                "resolution_notes":{"label":"הערות סגירה","type":"text", "filter": ServiceCall.resolution_notes,"get": lambda o: o.resolution_notes},
            },
            "filter_joins": {Elevator: (ServiceCall.elevator_id == Elevator.id)},
        },
        "elevators": {
            "label_he": "מעליות",
            "model": Elevator,
            "eager": [joinedload(Elevator.management_company), joinedload(Elevator.customer)],
            "default_columns": ["address", "city", "status", "model", "manufacturer", "contract_end", "risk_score"],
            "columns": {
                "id":              {"label_he": "מזהה",        "type": "text",   "filter": Elevator.id,              "get": lambda o: str(o.id)},
                "internal_number": {"label_he": "מס' פנימי",   "type": "text",   "filter": Elevator.internal_number, "get": lambda o: o.internal_number},
                "address":         {"label_he": "כתובת",       "type": "text",   "filter": Elevator.address,         "get": lambda o: o.address},
                "city":            {"label_he": "עיר",         "type": "text",   "filter": Elevator.city,            "get": lambda o: o.city},
                "status":          {"label_he": "סטטוס",       "type": "select", "filter": Elevator.status,          "get": lambda o: o.status,
                                    "options": ["ACTIVE","INACTIVE","UNDER_REPAIR"]},
                "model":           {"label_he": "דגם",         "type": "text",   "filter": Elevator.model,           "get": lambda o: o.model},
                "manufacturer":    {"label_he": "יצרן",        "type": "text",   "filter": Elevator.manufacturer,    "get": lambda o: o.manufacturer},
                "floor_count":     {"label_he": "קומות",       "type": "number", "filter": Elevator.floor_count,     "get": lambda o: o.floor_count},
                "contract_end":    {"label_he": "סיום חוזה",  "type": "date",   "filter": Elevator.contract_end,    "get": lambda o: _fmt(o.contract_end)},
                "risk_score":      {"label_he": "ציון סיכון",  "type": "number", "filter": Elevator.risk_score,      "get": lambda o: round(o.risk_score or 0, 1)},
                "has_debt":        {"label_he": "חוב",         "type": "bool",   "filter": Elevator.has_debt,        "get": lambda o: _fmt(o.has_debt)},
                "warranty_end":    {"label_he": "אחריות עד",   "type": "date",   "filter": Elevator.warranty_end,    "get": lambda o: _fmt(o.warranty_end)},
                "management_company": {"label_he": "חברת ניהול","type": "text",  "filter": None,
                                       "get": lambda o: _safe(o, "management_company", "name")},
                "customer":        {"label_he": "לקוח",        "type": "text",   "filter": None,
                                    "get": lambda o: _safe(o, "customer", "name")},
                "last_service_date":{"label":"שירות אחרון", "type": "date",   "filter": Elevator.last_service_date,"get": lambda o: _fmt(o.last_service_date)},
                "next_service_date":{"label":"שירות הבא",   "type": "date",   "filter": Elevator.next_service_date,"get": lambda o: _fmt(o.next_service_date)},
            },
            "filter_joins": {},
        },
        "customers": {
            "label_he": "לקוחות",
            "model": Customer,
            "eager": [],
            "default_columns": ["name", "customer_type", "city", "phone", "email", "is_active"],
            "columns": {
                "id":             {"label_he": "מזהה",       "type": "text",   "filter": Customer.id,             "get": lambda o: str(o.id)},
                "name":           {"label_he": "שם",         "type": "text",   "filter": Customer.name,           "get": lambda o: o.name},
                "customer_type":  {"label_he": "סוג",        "type": "select", "filter": Customer.customer_type,  "get": lambda o: o.customer_type,
                                   "options": ["PRIVATE","BUSINESS","MUNICIPALITY","MANAGEMENT_COMPANY","COMMITTEE"]},
                "city":           {"label_he": "עיר",        "type": "text",   "filter": Customer.city,           "get": lambda o: o.city},
                "phone":          {"label_he": "טלפון",      "type": "text",   "filter": Customer.phone,          "get": lambda o: o.phone},
                "email":          {"label_he": "אימייל",     "type": "text",   "filter": Customer.email,          "get": lambda o: o.email},
                "contact_person": {"label_he": "איש קשר",   "type": "text",   "filter": Customer.contact_person,  "get": lambda o: o.contact_person},
                "vat_number":     {"label_he": "ח.פ",        "type": "text",   "filter": Customer.vat_number,     "get": lambda o: o.vat_number},
                "payment_terms":  {"label_he": "תנאי תשלום", "type": "number", "filter": Customer.payment_terms,  "get": lambda o: o.payment_terms},
                "is_active":      {"label_he": "פעיל",       "type": "bool",   "filter": Customer.is_active,      "get": lambda o: _fmt(o.is_active)},
                "created_at":     {"label_he": "תאריך הצטרפות","type":"datetime","filter": Customer.created_at,   "get": lambda o: _fmt(o.created_at)},
            },
            "filter_joins": {},
        },
        "invoices": {
            "label_he": "חשבוניות",
            "model": Invoice,
            "eager": [joinedload(Invoice.customer)],
            "default_columns": ["number", "customer_name", "total", "status", "issue_date", "due_date"],
            "columns": {
                "id":            {"label_he": "מזהה",      "type": "text",   "filter": Invoice.id,           "get": lambda o: str(o.id)},
                "number":        {"label_he": "מספר",      "type": "text",   "filter": Invoice.number,       "get": lambda o: o.number},
                "customer_name": {"label_he": "לקוח",      "type": "text",   "filter": None,
                                  "get": lambda o: _safe(o, "customer", "name")},
                "total":         {"label_he": "סכום",      "type": "number", "filter": Invoice.total,        "get": lambda o: float(o.total or 0)},
                "amount_paid":   {"label_he": "שולם",      "type": "number", "filter": Invoice.amount_paid,  "get": lambda o: float(o.amount_paid or 0)},
                "status":        {"label_he": "סטטוס",     "type": "select", "filter": Invoice.status,       "get": lambda o: o.status,
                                  "options": ["DRAFT","SENT","PARTIAL","PAID","OVERDUE","CANCELLED"]},
                "issue_date":    {"label_he": "תאריך הנפקה","type": "date",  "filter": Invoice.issue_date,   "get": lambda o: _fmt(o.issue_date)},
                "due_date":      {"label_he": "תאריך פירעון","type":"date",  "filter": Invoice.due_date,     "get": lambda o: _fmt(o.due_date)},
                "paid_at":       {"label_he": "שולם בתאריך","type":"datetime","filter": Invoice.paid_at,     "get": lambda o: _fmt(o.paid_at)},
                "invoice_type":  {"label_he": "סוג",       "type": "select", "filter": Invoice.invoice_type, "get": lambda o: o.invoice_type,
                                  "options": ["TAX","RECEIPT","PROFORMA","CREDIT"]},
            },
            "filter_joins": {},
        },
        "inventory": {
            "label_he": "מלאי",
            "model": Part,
            "eager": [],
            "default_columns": ["name", "sku", "category", "quantity", "min_quantity", "unit_cost"],
            "columns": {
                "id":           {"label_he": "מזהה",     "type": "text",   "filter": Part.id,           "get": lambda o: str(o.id)},
                "sku":          {"label_he": "מק\"ט",    "type": "text",   "filter": Part.sku,          "get": lambda o: o.sku},
                "name":         {"label_he": "שם",       "type": "text",   "filter": Part.name,         "get": lambda o: o.name},
                "category":     {"label_he": "קטגוריה",  "type": "text",   "filter": Part.category,     "get": lambda o: o.category},
                "quantity":     {"label_he": "כמות",     "type": "number", "filter": Part.quantity,     "get": lambda o: o.quantity},
                "min_quantity": {"label_he": "מינימום",  "type": "number", "filter": Part.min_quantity, "get": lambda o: o.min_quantity},
                "unit_cost":    {"label_he": "עלות יחידה","type": "number","filter": Part.unit_cost,    "get": lambda o: float(o.unit_cost or 0)},
                "unit_price":   {"label_he": "מחיר מכירה","type": "number","filter": Part.unit_price,   "get": lambda o: float(o.unit_price or 0)},
                "location":     {"label_he": "מיקום",    "type": "text",   "filter": Part.location,     "get": lambda o: o.location},
                "is_active":    {"label_he": "פעיל",     "type": "bool",   "filter": Part.is_active,    "get": lambda o: _fmt(o.is_active)},
            },
            "filter_joins": {},
        },
        "maintenance": {
            "label_he": "תחזוקה",
            "model": MaintenanceSchedule,
            "eager": [joinedload(MaintenanceSchedule.elevator), joinedload(MaintenanceSchedule.technician)],
            "default_columns": ["scheduled_date", "address", "city", "maintenance_type", "status", "technician_name"],
            "columns": {
                "id":               {"label_he": "מזהה",     "type": "text",   "filter": MaintenanceSchedule.id,               "get": lambda o: str(o.id)},
                "scheduled_date":   {"label_he": "תאריך",    "type": "date",   "filter": MaintenanceSchedule.scheduled_date,   "get": lambda o: _fmt(o.scheduled_date)},
                "maintenance_type": {"label_he": "סוג",      "type": "select", "filter": MaintenanceSchedule.maintenance_type, "get": lambda o: o.maintenance_type,
                                     "options": ["QUARTERLY","SEMI_ANNUAL","ANNUAL","INSPECTION"]},
                "status":           {"label_he": "סטטוס",    "type": "select", "filter": MaintenanceSchedule.status,           "get": lambda o: o.status,
                                     "options": ["SCHEDULED","COMPLETED","OVERDUE","CANCELLED"]},
                "address":          {"label_he": "כתובת",    "type": "text",   "filter": Elevator.address,                     "get": lambda o: _safe(o, "elevator", "address")},
                "city":             {"label_he": "עיר",      "type": "text",   "filter": Elevator.city,                        "get": lambda o: _safe(o, "elevator", "city")},
                "technician_name":  {"label_he": "טכנאי",    "type": "text",   "filter": Technician.name,                      "get": lambda o: _safe(o, "technician", "name")},
                "completed_at":     {"label_he": "הושלם",    "type": "datetime","filter": MaintenanceSchedule.completed_at,    "get": lambda o: _fmt(o.completed_at)},
                "completion_notes": {"label_he": "הערות",    "type": "text",   "filter": MaintenanceSchedule.completion_notes, "get": lambda o: o.completion_notes},
            },
            "filter_joins": {
                Elevator:   (MaintenanceSchedule.elevator_id   == Elevator.id),
                Technician: (MaintenanceSchedule.technician_id == Technician.id),
            },
        },
        "contracts": {
            "label_he": "חוזים",
            "model": Contract,
            "eager": [joinedload(Contract.customer)],
            "default_columns": ["number", "customer_name", "contract_type", "status", "start_date", "end_date", "monthly_value"],
            "columns": {
                "id":            {"label_he": "מזהה",      "type": "text",   "filter": Contract.id,           "get": lambda o: str(o.id)},
                "number":        {"label_he": "מספר",      "type": "text",   "filter": Contract.number,       "get": lambda o: o.number},
                "customer_name": {"label_he": "לקוח",      "type": "text",   "filter": None,
                                  "get": lambda o: _safe(o, "customer", "name")},
                "contract_type": {"label_he": "סוג",       "type": "select", "filter": Contract.contract_type,"get": lambda o: o.contract_type,
                                  "options": ["MAINTENANCE","INSPECTION","COMPREHENSIVE","OTHER"]},
                "status":        {"label_he": "סטטוס",     "type": "select", "filter": Contract.status,       "get": lambda o: o.status,
                                  "options": ["ACTIVE","EXPIRED","CANCELLED","DRAFT"]},
                "start_date":    {"label_he": "התחלה",     "type": "date",   "filter": Contract.start_date,   "get": lambda o: _fmt(o.start_date)},
                "end_date":      {"label_he": "סיום",      "type": "date",   "filter": Contract.end_date,     "get": lambda o: _fmt(o.end_date)},
                "monthly_value": {"label_he": "תשלום חודשי","type":"number", "filter": Contract.monthly_value,"get": lambda o: float(o.monthly_value or 0)},
                "auto_renew":    {"label_he": "חידוש אוטו","type": "bool",   "filter": Contract.auto_renew,   "get": lambda o: _fmt(o.auto_renew)},
            },
            "filter_joins": {},
        },
        "leads": {
            "label_he": "לידים",
            "model": Lead,
            "eager": [joinedload(Lead.customer)],
            "default_columns": ["created_at", "customer_name", "source", "status", "estimated_value"],
            "columns": {
                "id":              {"label_he": "מזהה",    "type": "text",   "filter": Lead.id,              "get": lambda o: str(o.id)},
                "customer_name":   {"label_he": "לקוח",    "type": "text",   "filter": None,
                                    "get": lambda o: _safe(o, "customer", "name")},
                "source":          {"label_he": "מקור",    "type": "text",   "filter": Lead.source,          "get": lambda o: o.source},
                "status":          {"label_he": "סטטוס",   "type": "select", "filter": Lead.status,          "get": lambda o: o.status,
                                    "options": ["NEW","CONTACTED","QUALIFIED","PROPOSAL","WON","LOST"]},
                "estimated_value": {"label_he": "שווי משוער","type":"number","filter": Lead.estimated_value, "get": lambda o: float(o.estimated_value or 0)},
                "created_at":      {"label_he": "נוצר",    "type": "datetime","filter": Lead.created_at,     "get": lambda o: _fmt(o.created_at)},
                "notes":           {"label_he": "הערות",   "type": "text",   "filter": Lead.notes,           "get": lambda o: o.notes},
            },
            "filter_joins": {},
        },
        "inspections": {
            "label_he": "ביקורות",
            "model": InspectionReport,
            "eager": [],
            "default_columns": ["inspection_date", "raw_address", "raw_city", "result", "deficiency_count", "report_status"],
            "columns": {
                "id":               {"label_he": "מזהה",     "type": "text",   "filter": InspectionReport.id,              "get": lambda o: str(o.id)},
                "inspection_date":  {"label_he": "תאריך",    "type": "date",   "filter": InspectionReport.inspection_date, "get": lambda o: _fmt(o.inspection_date)},
                "raw_address":      {"label_he": "כתובת",    "type": "text",   "filter": InspectionReport.raw_address,     "get": lambda o: o.raw_address},
                "raw_city":         {"label_he": "עיר",      "type": "text",   "filter": InspectionReport.raw_city,        "get": lambda o: o.raw_city},
                "result":           {"label_he": "תוצאה",    "type": "select", "filter": InspectionReport.result,         "get": lambda o: o.result,
                                     "options": ["PASS","FAIL","UNKNOWN"]},
                "inspector_name":   {"label_he": "בודק",     "type": "text",   "filter": InspectionReport.inspector_name,  "get": lambda o: o.inspector_name},
                "deficiency_count": {"label_he": "ליקויים",  "type": "number", "filter": InspectionReport.deficiency_count,"get": lambda o: o.deficiency_count},
                "report_status":    {"label_he": "סטטוס טיפול","type":"select","filter": InspectionReport.report_status,  "get": lambda o: o.report_status,
                                     "options": ["NA","OPEN","PARTIAL","CLOSED"]},
                "match_status":     {"label_he": "התאמה",    "type": "select", "filter": InspectionReport.match_status,   "get": lambda o: o.match_status},
            },
            "filter_joins": {},
        },
    }


_SCHEMAS: Optional[dict] = None


def get_schemas() -> dict:
    global _SCHEMAS
    if _SCHEMAS is None:
        _SCHEMAS = _build_schemas()
    return _SCHEMAS


# ---------------------------------------------------------------------------
# Filter application
# ---------------------------------------------------------------------------

def _apply_filter(query, attr, op: str, value: Any, col_type: str):
    if attr is None:
        return query

    # Type coercion
    if col_type == "number":
        try:
            value = float(value)
        except (TypeError, ValueError):
            return query
    elif col_type in ("date", "datetime") and isinstance(value, str):
        try:
            value = datetime.fromisoformat(value.replace("Z", "+00:00"))
        except ValueError:
            try:
                value = date.fromisoformat(value)
            except ValueError:
                return query
    elif col_type == "bool":
        if isinstance(value, str):
            value = value.lower() in ("true", "1", "yes", "כן")

    if op == "eq":
        return query.filter(attr == value)
    elif op == "neq":
        return query.filter(attr != value)
    elif op == "contains":
        return query.filter(attr.ilike(f"%{value}%"))
    elif op == "starts_with":
        return query.filter(attr.ilike(f"{value}%"))
    elif op == "gt":
        return query.filter(attr > value)
    elif op == "gte":
        return query.filter(attr >= value)
    elif op == "lt":
        return query.filter(attr < value)
    elif op == "lte":
        return query.filter(attr <= value)
    elif op == "in":
        vals = value if isinstance(value, list) else [value]
        return query.filter(attr.in_(vals))
    elif op == "is_null":
        return query.filter(attr == None)  # noqa: E711
    elif op == "is_not_null":
        return query.filter(attr != None)  # noqa: E711
    return query


# ---------------------------------------------------------------------------
# Main query function
# ---------------------------------------------------------------------------

def run_report(
    db: Session,
    entity_type: str,
    columns: list[str],
    filters: list[dict],
    sort_by: Optional[str],
    sort_dir: str,
    skip: int,
    limit: int,
    include_custom_fields: bool = True,
) -> dict:
    schemas = get_schemas()
    schema = schemas.get(entity_type)
    if not schema:
        raise ValueError(f"Unknown entity type: {entity_type}")

    col_defs = schema["columns"]
    model = schema["model"]

    # Ensure "id" is always fetched (needed for custom field lookup)
    fetch_cols = list(dict.fromkeys(["id"] + [c for c in columns if c in col_defs]))

    # Base query with eager loads
    query = db.query(model)
    for opt in schema.get("eager", []):
        query = query.options(opt)

    # Joins needed for filters on related models
    joined = set()
    filter_joins = schema.get("filter_joins", {})
    for f in filters:
        field_key = f.get("field", "")
        if field_key not in col_defs:
            continue
        fattr = col_defs[field_key].get("filter")
        if fattr is None:
            continue
        # Check if fattr belongs to a join model
        for join_model, join_cond in filter_joins.items():
            if hasattr(fattr, "class_") and fattr.class_ is join_model and join_model not in joined:
                query = query.outerjoin(join_model, join_cond)
                joined.add(join_model)
                break

    # Apply filters
    for f in filters:
        field_key = f.get("field", "")
        if field_key not in col_defs:
            continue
        fattr = col_defs[field_key].get("filter")
        col_type = col_defs[field_key].get("type", "text")
        query = _apply_filter(query, fattr, f.get("op", "eq"), f.get("value"), col_type)

    total = query.count()

    # Sort
    if sort_by and sort_by in col_defs:
        sort_attr = col_defs[sort_by].get("filter")
        if sort_attr is not None:
            query = query.order_by(sort_attr.desc() if sort_dir == "desc" else sort_attr.asc())

    objects = query.offset(skip).limit(min(limit, 1000)).all()

    # Extract rows
    rows = []
    entity_ids = []
    for obj in objects:
        row = {}
        for col_key in fetch_cols:
            try:
                row[col_key] = col_defs[col_key]["get"](obj)
            except Exception:
                row[col_key] = None
        rows.append(row)
        if row.get("id"):
            try:
                entity_ids.append(uuid.UUID(str(row["id"])))
            except Exception:
                pass

    # Merge custom field values
    if include_custom_fields and entity_ids:
        from app.models.custom_field import CustomFieldValue, CustomField
        cfv_rows = (
            db.query(CustomFieldValue, CustomField.field_name, CustomField.field_label)
            .join(CustomField, CustomFieldValue.field_id == CustomField.id)
            .filter(
                CustomFieldValue.entity_type == entity_type,
                CustomFieldValue.entity_id.in_(entity_ids),
                CustomField.is_active == True,
            )
            .all()
        )
        cf_map: dict[str, dict] = {}
        cf_labels: dict[str, str] = {}
        for cfv, fname, flabel in cfv_rows:
            key = str(cfv.entity_id)
            if key not in cf_map:
                cf_map[key] = {}
            cf_map[key][f"cf_{fname}"] = cfv.value
            cf_labels[f"cf_{fname}"] = flabel

        for row in rows:
            eid = str(row.get("id", ""))
            row.update(cf_map.get(eid, {}))
    else:
        cf_labels = {}

    # Columns metadata for frontend
    columns_meta = [
        {
            "key": c,
            "label_he": col_defs[c]["label_he"] if c in col_defs else cf_labels.get(c, c),
            "type": col_defs[c].get("type", "text") if c in col_defs else "text",
        }
        for c in fetch_cols
    ]

    # Remove "id" from output if not requested by user (but keep it for CF lookup)
    if "id" not in columns:
        for row in rows:
            row.pop("id", None)
        columns_meta = [m for m in columns_meta if m["key"] != "id"]

    return {"total": total, "rows": rows, "columns_meta": columns_meta}


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_to_excel(result: dict, entity_label: str) -> io.BytesIO:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = entity_label[:31]

    headers = [m["label_he"] for m in result["columns_meta"]]
    keys = [m["key"] for m in result["columns_meta"]]

    header_fill = PatternFill("solid", fgColor="1a2744")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="right")
        ws.column_dimensions[cell.column_letter].width = max(15, len(header) + 2)

    for row_idx, row in enumerate(result["rows"], 2):
        for col_idx, key in enumerate(keys, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
