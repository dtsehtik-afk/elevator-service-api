"""Analytics queries — aggregated statistics for management reporting."""

import calendar
import io
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List

from sqlalchemy.orm import Session

from app.models.assignment import Assignment
from app.models.elevator import Elevator
from app.models.service_call import ServiceCall
from app.models.technician import Technician


def get_recurring_fault_elevators(db: Session) -> List[Dict[str, Any]]:
    """Return elevators with 3 or more service calls in the last 90 days.

    Args:
        db: Database session.

    Returns:
        List of dicts: elevator info + call count.
    """
    ninety_days_ago = datetime.now(timezone.utc) - timedelta(days=90)

    # Fetch all recent calls and aggregate in Python to stay DB-agnostic
    recent_calls = (
        db.query(ServiceCall)
        .filter(ServiceCall.created_at >= ninety_days_ago)
        .all()
    )

    # Aggregate per elevator
    elevator_stats: Dict[Any, Dict[str, int]] = {}
    for call in recent_calls:
        eid = call.elevator_id
        if eid not in elevator_stats:
            elevator_stats[eid] = {"total": 0, "recurring": 0}
        elevator_stats[eid]["total"] += 1
        if call.is_recurring:
            elevator_stats[eid]["recurring"] += 1

    # Filter to >= 3 calls
    eligible_ids = [eid for eid, stats in elevator_stats.items() if stats["total"] >= 3]
    if not eligible_ids:
        return []

    elevators = (
        db.query(Elevator)
        .filter(Elevator.id.in_(eligible_ids))
        .all()
    )

    results = [
        {
            "elevator_id": str(e.id),
            "address": e.address,
            "city": e.city,
            "building_name": e.building_name,
            "risk_score": e.risk_score,
            "call_count_90_days": elevator_stats[e.id]["total"],
            "recurring_count": elevator_stats[e.id]["recurring"],
        }
        for e in elevators
    ]
    return sorted(results, key=lambda x: x["call_count_90_days"], reverse=True)


def get_technician_performance(db: Session) -> List[Dict[str, Any]]:
    """Compute average resolution time and assignment counts per technician.

    Args:
        db: Database session.

    Returns:
        List of dicts with performance metrics per technician.
    """
    technicians = db.query(Technician).filter(Technician.is_active == True).all()  # noqa: E712
    results: List[Dict[str, Any]] = []

    for tech in technicians:
        assignments = (
            db.query(Assignment)
            .filter(Assignment.technician_id == tech.id)
            .all()
        )
        call_ids = [a.service_call_id for a in assignments]

        if not call_ids:
            results.append(
                {
                    "technician_id": str(tech.id),
                    "name": tech.name,
                    "email": tech.email,
                    "total_assigned": 0,
                    "total_resolved": 0,
                    "avg_resolution_hours": None,
                }
            )
            continue

        calls = db.query(ServiceCall).filter(ServiceCall.id.in_(call_ids)).all()
        resolved = [c for c in calls if c.resolved_at and c.created_at]
        total_hours = sum(
            (c.resolved_at - c.created_at).total_seconds() / 3600 for c in resolved
        )
        avg = round(total_hours / len(resolved), 2) if resolved else None

        results.append(
            {
                "technician_id": str(tech.id),
                "name": tech.name,
                "email": tech.email,
                "total_assigned": len(assignments),
                "total_resolved": len(resolved),
                "avg_resolution_hours": avg,
            }
        )

    return sorted(results, key=lambda x: x["total_assigned"], reverse=True)


def get_monthly_summary(db: Session, year: int, month: int) -> Dict[str, Any]:
    """Return aggregated service call statistics for a specific month.

    Args:
        db: Database session.
        year: The year (e.g. 2024).
        month: The month number (1-12).

    Returns:
        Summary dict with totals, averages, and breakdowns.
    """
    first_day = datetime(year, month, 1, tzinfo=timezone.utc)
    last_day_num = calendar.monthrange(year, month)[1]
    last_day = datetime(year, month, last_day_num, 23, 59, 59, tzinfo=timezone.utc)

    calls = (
        db.query(ServiceCall)
        .filter(
            ServiceCall.created_at >= first_day,
            ServiceCall.created_at <= last_day,
        )
        .all()
    )

    total = len(calls)
    resolved = [c for c in calls if c.status in ("RESOLVED", "CLOSED")]
    recurring = [c for c in calls if c.is_recurring]

    by_priority: Dict[str, int] = {}
    by_fault: Dict[str, int] = {}
    for c in calls:
        by_priority[c.priority] = by_priority.get(c.priority, 0) + 1
        by_fault[c.fault_type] = by_fault.get(c.fault_type, 0) + 1

    total_hours = sum(
        (c.resolved_at - c.created_at).total_seconds() / 3600
        for c in resolved
        if c.resolved_at and c.created_at
    )
    avg_hours = round(total_hours / len(resolved), 2) if resolved else None

    return {
        "year": year,
        "month": month,
        "total_calls": total,
        "resolved_calls": len(resolved),
        "recurring_calls": len(recurring),
        "resolution_rate": round(len(resolved) / total * 100, 1) if total else 0,
        "avg_resolution_hours": avg_hours,
        "calls_by_priority": by_priority,
        "calls_by_fault_type": by_fault,
    }


def get_elevator_history(db: Session, elevator_id) -> Dict[str, Any]:
    """Return full service call history for a specific elevator."""
    elevator = db.query(Elevator).filter(Elevator.id == elevator_id).first()
    if not elevator:
        return {}

    calls = (
        db.query(ServiceCall)
        .filter(ServiceCall.elevator_id == elevator_id)
        .order_by(ServiceCall.created_at.desc())
        .all()
    )

    call_list = []
    for c in calls:
        # Find assigned technician
        assignment = (
            db.query(Assignment)
            .filter(Assignment.service_call_id == c.id,
                    Assignment.status.in_(["CONFIRMED", "COMPLETED"]))
            .order_by(Assignment.assigned_at.desc())
            .first()
        )
        tech_name = None
        if assignment:
            tech = db.query(Technician).filter(Technician.id == assignment.technician_id).first()
            tech_name = tech.name if tech else None

        resolution_hours = None
        if c.resolved_at and c.created_at:
            resolution_hours = round((c.resolved_at - c.created_at).total_seconds() / 3600, 1)

        call_list.append({
            "id": str(c.id),
            "created_at": c.created_at.isoformat() if c.created_at else None,
            "resolved_at": c.resolved_at.isoformat() if c.resolved_at else None,
            "status": c.status,
            "fault_type": c.fault_type,
            "priority": c.priority,
            "description": c.description,
            "reported_by": c.reported_by,
            "technician": tech_name,
            "resolution_hours": resolution_hours,
            "resolution_notes": c.resolution_notes,
            "is_recurring": c.is_recurring,
        })

    return {
        "elevator_id": str(elevator.id),
        "address": elevator.address,
        "city": elevator.city,
        "building_name": elevator.building_name,
        "serial_number": elevator.serial_number,
        "total_calls": len(calls),
        "calls": call_list,
    }


def export_calls_excel(
    db: Session,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    elevator_id=None,
) -> bytes:
    """
    Export service calls to an Excel file.
    Returns raw bytes of the .xlsx file.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl not installed — run: pip install openpyxl")

    query = db.query(ServiceCall)
    if date_from:
        query = query.filter(ServiceCall.created_at >= date_from)
    if date_to:
        query = query.filter(ServiceCall.created_at <= date_to)
    if elevator_id:
        query = query.filter(ServiceCall.elevator_id == elevator_id)

    calls = query.order_by(ServiceCall.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "קריאות שירות"
    ws.sheet_view.rightToLeft = True

    # Header style
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    headers = [
        "תאריך פתיחה", "עיר", "כתובת", "בניין", "סוג תקלה",
        "עדיפות", "סטטוס", "מדווח", "טכנאי", "שעות טיפול",
        "הערות סגירה", "חוזרת",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    _STATUS_HE = {
        "OPEN": "פתוח", "ASSIGNED": "שובץ", "IN_PROGRESS": "בטיפול",
        "RESOLVED": "טופל", "CLOSED": "סגור",
    }
    _FAULT_HE = {
        "STUCK": "מעלית תקועה", "DOOR": "תקלת דלת", "ELECTRICAL": "חשמלית",
        "MECHANICAL": "מכנית", "SOFTWARE": "תוכנה", "OTHER": "כללית",
    }
    _PRI_HE = {
        "CRITICAL": "קריטי", "HIGH": "גבוה", "MEDIUM": "בינוני", "LOW": "נמוך",
    }

    for row_num, c in enumerate(calls, 2):
        elevator = db.query(Elevator).filter(Elevator.id == c.elevator_id).first()
        assignment = (
            db.query(Assignment)
            .filter(Assignment.service_call_id == c.id,
                    Assignment.status.in_(["CONFIRMED", "COMPLETED"]))
            .first()
        )
        tech_name = ""
        if assignment:
            tech = db.query(Technician).filter(Technician.id == assignment.technician_id).first()
            tech_name = tech.name if tech else ""

        resolution_hours = ""
        if c.resolved_at and c.created_at:
            resolution_hours = round((c.resolved_at - c.created_at).total_seconds() / 3600, 1)

        row = [
            c.created_at.strftime("%d/%m/%Y %H:%M") if c.created_at else "",
            elevator.city if elevator else "",
            elevator.address if elevator else "",
            elevator.building_name if elevator else "",
            _FAULT_HE.get(c.fault_type, c.fault_type),
            _PRI_HE.get(c.priority, c.priority),
            _STATUS_HE.get(c.status, c.status),
            c.reported_by or "",
            tech_name,
            resolution_hours,
            c.resolution_notes or "",
            "כן" if c.is_recurring else "לא",
        ]
        for col, val in enumerate(row, 1):
            ws.cell(row=row_num, column=col, value=val)

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def get_risk_elevators(db: Session, threshold: float = 70.0) -> List[Dict[str, Any]]:
    """Return elevators with risk_score above the given threshold.

    Args:
        db: Database session.
        threshold: Minimum risk score (default 70).

    Returns:
        List of elevator dicts sorted by risk_score descending.
    """
    elevators = (
        db.query(Elevator)
        .filter(Elevator.risk_score > threshold, Elevator.status == "ACTIVE")
        .order_by(Elevator.risk_score.desc())
        .all()
    )

    return [
        {
            "elevator_id": str(e.id),
            "address": e.address,
            "city": e.city,
            "building_name": e.building_name,
            "status": e.status,
            "risk_score": e.risk_score,
            "last_service_date": e.last_service_date.isoformat() if e.last_service_date else None,
            "next_service_date": e.next_service_date.isoformat() if e.next_service_date else None,
        }
        for e in elevators
    ]
